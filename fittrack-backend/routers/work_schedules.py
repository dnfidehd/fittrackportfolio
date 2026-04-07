from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import get_db
from models import WorkSchedule, WorkScheduleTemplate, Member, Gym
from schemas import (
    WorkScheduleCreate, WorkScheduleUpdate, WorkScheduleResponse,
    WorkScheduleTemplateCreate, WorkScheduleTemplateResponse,
    WorkScheduleBulkCreate
)
from routers.auth import get_current_user
from constants import Role
from utils.auth import assert_roles

router = APIRouter(
    tags=["work_schedules"],
    responses={404: {"description": "Not found"}},
)


READABLE_WORK_SCHEDULE_ROLES = [Role.COACH, Role.SUBCOACH, Role.USER]
EDITABLE_WORK_SCHEDULE_ROLES = [Role.COACH, Role.SUBCOACH]


def assert_work_schedule_read_access(current_user: Member) -> None:
    assert_roles(current_user, READABLE_WORK_SCHEDULE_ROLES)


def assert_work_schedule_edit_access(current_user: Member) -> None:
    assert_roles(current_user, EDITABLE_WORK_SCHEDULE_ROLES)


def resolve_accessible_gym_id(
    db: Session,
    current_user: Member,
    requested_gym_id: Optional[int],
) -> int:
    if requested_gym_id is None:
        return current_user.gym_id or 1

    coach_in_gym = db.query(Member).filter(
        Member.phone == current_user.phone,
        Member.gym_id == requested_gym_id,
        Member.role.in_([Role.COACH, Role.SUBCOACH]),
    ).first()

    if not coach_in_gym and current_user.role == Role.USER:
        raise HTTPException(status_code=403, detail="해당 지점에 접근할 권한이 없습니다")

    return requested_gym_id


def build_schedule_response(schedule: WorkSchedule, coach_name: str) -> WorkScheduleResponse:
    """근무 일정 응답 객체 생성"""
    return WorkScheduleResponse(
        id=schedule.id,
        gym_id=schedule.gym_id,
        coach_id=schedule.coach_id,
        coach_name=coach_name,
        date=schedule.date,
        start_time=schedule.start_time,
        end_time=schedule.end_time,
        work_category=getattr(schedule, 'work_category', 'general'),
        shift_type=schedule.shift_type,
        memo=schedule.memo,
        status=schedule.status,
        created_at=schedule.created_at,
        updated_at=getattr(schedule, 'updated_at', schedule.created_at)
    )


def build_template_response(template: WorkScheduleTemplate, coach_name: str) -> WorkScheduleTemplateResponse:
    """근무 템플릿 응답 객체 생성"""
    return WorkScheduleTemplateResponse(
        id=template.id,
        gym_id=template.gym_id,
        coach_id=template.coach_id,
        coach_name=coach_name,
        start_time=template.start_time,
        end_time=template.end_time,
        shift_type=template.shift_type,
        days_of_week=template.days_of_week,
        memo=template.memo,
        created_at=template.created_at
    )


# ✅ 1. 월별 근무표 조회 (캘린더 그리드용)
@router.get("/", response_model=List[WorkScheduleResponse])
def get_monthly_work_schedules(
    year_month: str = Query(..., description="YYYY-MM 형식"),
    coach_id: Optional[int] = None,
    gym_id: Optional[int] = None,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    월별 근무 일정 조회
    - year_month: "2026-03"
    - coach_id: 특정 코치만 필터링 (선택사항)
    - gym_id: 특정 지점만 필터링 (선택사항, 멀티박스)
    """
    assert_work_schedule_read_access(current_user)
    gym_id = resolve_accessible_gym_id(db, current_user, gym_id)

    # year_month 파싱
    try:
        year, month = map(int, year_month.split('-'))
        # 해당 월의 첫날과 마지막날
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM")

    # 기본 쿼리
    query = db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.date >= first_day,
        WorkSchedule.date <= last_day
    )

    # coach_id 필터 (선택사항)
    if coach_id:
        query = query.filter(WorkSchedule.coach_id == coach_id)

    schedules = query.order_by(WorkSchedule.date, WorkSchedule.start_time).all()

    # 템플릿 기반 자동 생성 (누락된 근무 추가)
    templates = db.query(WorkScheduleTemplate).filter(
        WorkScheduleTemplate.gym_id == gym_id
    ).all()

    # 기존 근무의 (date, coach_id, start_time) 셋
    existing_keys = {(s.date, s.coach_id, s.start_time) for s in schedules}

    # 템플릿 기반 근무 생성
    new_schedules = []
    for tmpl in templates:
        # coach_id 필터가 있으면 해당 코치만 처리
        if coach_id and tmpl.coach_id != coach_id:
            continue

        # 요일 리스트
        days_of_week = [int(d) for d in tmpl.days_of_week.split(',')]

        # 해당 월의 모든 날짜 순회
        current = first_day
        while current <= last_day:
            weekday = current.weekday()  # 0:월, 1:화, ..., 6:일

            # 템플릿의 요일에 맞는지 확인
            if weekday in days_of_week:
                key = (current, tmpl.coach_id, tmpl.start_time)
                # 중복 체크
                if key not in existing_keys:
                    new_work = WorkSchedule(
                        gym_id=gym_id,
                        coach_id=tmpl.coach_id,
                        date=current,
                        start_time=tmpl.start_time,
                        end_time=tmpl.end_time,
                        shift_type=tmpl.shift_type,
                        memo=tmpl.memo,
                        status="scheduled"
                    )
                    db.add(new_work)
                    new_schedules.append(new_work)
                    existing_keys.add(key)

            current += timedelta(days=1)

    if new_schedules:
        db.commit()
        for ns in new_schedules:
            db.refresh(ns)
        # 전체 다시 조회
        query = db.query(WorkSchedule).filter(
            WorkSchedule.gym_id == gym_id,
            WorkSchedule.date >= first_day,
            WorkSchedule.date <= last_day
        )
        if coach_id:
            query = query.filter(WorkSchedule.coach_id == coach_id)
        schedules = query.order_by(WorkSchedule.date, WorkSchedule.start_time).all()

    # Response 변환
    results = []
    for schedule in schedules:
        coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
        coach_name = coach.name if coach else "Unknown"
        response = build_schedule_response(schedule, coach_name)
        results.append(response)

    return results


# ✅ 2. 특정 날짜 근무 조회
@router.get("/daily", response_model=List[WorkScheduleResponse])
def get_daily_work_schedules(
    date_str: str = Query(..., description="YYYY-MM-DD 형식"),
    gym_id: Optional[int] = None,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    특정 날짜의 모든 근무 일정 조회
    """
    assert_work_schedule_read_access(current_user)
    gym_id = resolve_accessible_gym_id(db, current_user, gym_id)

    try:
        target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    schedules = db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.date == target_date
    ).order_by(WorkSchedule.start_time).all()

    # Response 변환
    results = []
    for schedule in schedules:
        coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
        coach_name = coach.name if coach else "Unknown"
        response = build_schedule_response(schedule, coach_name)
        results.append(response)

    return results


# ✅ 2-1. 특정 날짜 근무 일괄 등록 (Bulk Insert)
@router.post("/daily/bulk", response_model=List[WorkScheduleResponse])
def create_bulk_work_schedules(
    data: WorkScheduleBulkCreate,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    특정 날짜의 근무 일정을 한 번에 일괄 업데이트 (기존 내역 삭제 후 덮어쓰기)
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1
    target_date = data.date

    # 1. 해당 날짜, 해당 지점의 모든 근무 일정 삭제 (완전 덮어쓰기 기준)
    db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.date == target_date
    ).delete()

    # 2. 새 일정들 일괄 생성
    new_schedules = []
    for item in data.schedules:
        # 시간 유효성 검사 생략 (프론트에서 처리 가정, 혹은 간단히만 방어)
        work_schedule = WorkSchedule(
            gym_id=gym_id,
            coach_id=item.coach_id,
            date=item.date,
            start_time=item.start_time,
            end_time=item.end_time,
            work_category=item.work_category,
            shift_type=item.shift_type,
            memo=item.memo,
            status="scheduled"
        )
        db.add(work_schedule)
        new_schedules.append(work_schedule)
    
    db.commit()

    # Response 반환
    results = []
    for schedule in new_schedules:
        db.refresh(schedule)
        coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
        coach_name = coach.name if coach else "Unknown"
        results.append(build_schedule_response(schedule, coach_name))

    return results


# ✅ 3. 단일 근무 일정 생성 (기존)
@router.post("/", response_model=WorkScheduleResponse)
def create_work_schedule(
    data: WorkScheduleCreate,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    새로운 근무 일정 생성
    - 시간대 중복 검증 포함
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1

    # 시간대 유효성 검증
    try:
        start = datetime.strptime(data.start_time, "%H:%M").time()
        end = datetime.strptime(data.end_time, "%H:%M").time()
        if start >= end:
            raise HTTPException(status_code=400, detail="시작 시간이 종료 시간보다 늦을 수 없습니다")
    except ValueError:
        raise HTTPException(status_code=400, detail="시간 형식이 올바르지 않습니다 (HH:MM)")

    # 같은 날짜, 같은 코치의 기존 근무와 시간대 겹침 체크
    existing = db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.coach_id == data.coach_id,
        WorkSchedule.date == data.date
    ).all()

    for ex in existing:
        ex_start = datetime.strptime(ex.start_time, "%H:%M").time()
        ex_end = datetime.strptime(ex.end_time, "%H:%M").time()

        # 시간대 겹침 체크
        if not (end <= ex_start or start >= ex_end):
            raise HTTPException(
                status_code=400,
                detail=f"시간대 중복: {ex_start}-{ex_end}와 겹칩니다"
            )

    # 근무 생성
    work_schedule = WorkSchedule(
        gym_id=gym_id,
        coach_id=data.coach_id,
        date=data.date,
        start_time=data.start_time,
        end_time=data.end_time,
        shift_type=data.shift_type,
        memo=data.memo,
        status="scheduled"
    )

    db.add(work_schedule)
    db.commit()
    db.refresh(work_schedule)

    # Response 변환
    coach = db.query(Member).filter(Member.id == work_schedule.coach_id).first()
    coach_name = coach.name if coach else "Unknown"
    return build_schedule_response(work_schedule, coach_name)


# ✅ 4. 근무 일정 수정
@router.put("/{work_schedule_id}", response_model=WorkScheduleResponse)
def update_work_schedule(
    work_schedule_id: int,
    data: WorkScheduleUpdate,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    기존 근무 일정 수정
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1

    schedule = db.query(WorkSchedule).filter(
        WorkSchedule.id == work_schedule_id,
        WorkSchedule.gym_id == gym_id
    ).first()

    if not schedule:
        raise HTTPException(status_code=404, detail="근무 일정을 찾을 수 없습니다")

    # 시간 수정 시 중복 체크
    if data.start_time or data.end_time:
        new_start = data.start_time or schedule.start_time
        new_end = data.end_time or schedule.end_time

        try:
            start = datetime.strptime(new_start, "%H:%M").time()
            end = datetime.strptime(new_end, "%H:%M").time()
            if start >= end:
                raise HTTPException(status_code=400, detail="시작 시간이 종료 시간보다 늦을 수 없습니다")
        except ValueError:
            raise HTTPException(status_code=400, detail="시간 형식이 올바르지 않습니다 (HH:MM)")

        # 다른 근무와의 충돌 체크 (자신은 제외)
        existing = db.query(WorkSchedule).filter(
            WorkSchedule.gym_id == gym_id,
            WorkSchedule.coach_id == schedule.coach_id,
            WorkSchedule.date == schedule.date,
            WorkSchedule.id != work_schedule_id
        ).all()

        for ex in existing:
            ex_start = datetime.strptime(ex.start_time, "%H:%M").time()
            ex_end = datetime.strptime(ex.end_time, "%H:%M").time()

            if not (end <= ex_start or start >= ex_end):
                raise HTTPException(
                    status_code=400,
                    detail=f"시간대 중복: {ex_start}-{ex_end}와 겹칩니다"
                )

    # 필드 업데이트
    if data.start_time:
        schedule.start_time = data.start_time
    if data.end_time:
        schedule.end_time = data.end_time
    if data.shift_type:
        schedule.shift_type = data.shift_type
    if data.memo is not None:
        schedule.memo = data.memo
    if data.status:
        schedule.status = data.status

    schedule.updated_at = datetime.now()

    db.commit()
    db.refresh(schedule)

    # Response 변환
    coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
    coach_name = coach.name if coach else "Unknown"
    return build_schedule_response(schedule, coach_name)


# ✅ 5. 근무 일정 삭제
@router.delete("/{work_schedule_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_work_schedule(
    work_schedule_id: int,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    근무 일정 삭제
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1

    schedule = db.query(WorkSchedule).filter(
        WorkSchedule.id == work_schedule_id,
        WorkSchedule.gym_id == gym_id
    ).first()

    if not schedule:
        raise HTTPException(status_code=404, detail="근무 일정을 찾을 수 없습니다")

    db.delete(schedule)
    db.commit()

    return None


# ✅ 6. 반복 패턴 목록 조회
@router.get("/templates", response_model=List[WorkScheduleTemplateResponse])
def get_templates(
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    반복 근무 패턴 (고정 스케줄) 목록 조회
    """
    assert_work_schedule_read_access(current_user)

    gym_id = current_user.gym_id or 1

    templates = db.query(WorkScheduleTemplate).filter(
        WorkScheduleTemplate.gym_id == gym_id
    ).order_by(WorkScheduleTemplate.days_of_week).all()

    # Response 변환
    results = []
    for tmpl in templates:
        coach = db.query(Member).filter(Member.id == tmpl.coach_id).first()
        coach_name = coach.name if coach else "Unknown"
        response = build_template_response(tmpl, coach_name)
        results.append(response)

    return results


# ✅ 7. 반복 패턴 생성
@router.post("/templates", response_model=WorkScheduleTemplateResponse)
def create_template(
    data: WorkScheduleTemplateCreate,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    반복 근무 패턴 생성
    - days_of_week: "0,1,2,3,4" (월~금)
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1

    # 시간 유효성 검증
    try:
        start = datetime.strptime(data.start_time, "%H:%M").time()
        end = datetime.strptime(data.end_time, "%H:%M").time()
        if start >= end:
            raise HTTPException(status_code=400, detail="시작 시간이 종료 시간보다 늦을 수 없습니다")
    except ValueError:
        raise HTTPException(status_code=400, detail="시간 형식이 올바르지 않습니다 (HH:MM)")

    # 요일 유효성 검증
    try:
        days = [int(d) for d in data.days_of_week.split(',')]
        if any(d < 0 or d > 6 for d in days):
            raise ValueError
    except ValueError:
        raise HTTPException(status_code=400, detail="요일은 0-6 범위의 숫자여야 합니다 (쉼표 구분)")

    # 템플릿 생성
    template = WorkScheduleTemplate(
        gym_id=gym_id,
        coach_id=data.coach_id,
        start_time=data.start_time,
        end_time=data.end_time,
        shift_type=data.shift_type,
        days_of_week=data.days_of_week,
        memo=data.memo
    )

    db.add(template)
    db.commit()
    db.refresh(template)

    # Response 변환
    coach = db.query(Member).filter(Member.id == template.coach_id).first()
    coach_name = coach.name if coach else "Unknown"
    return build_template_response(template, coach_name)


# ✅ 8. 반복 패턴 삭제
@router.delete("/templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_template(
    template_id: int,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    반복 패턴 삭제
    """
    assert_work_schedule_edit_access(current_user)

    gym_id = current_user.gym_id or 1

    template = db.query(WorkScheduleTemplate).filter(
        WorkScheduleTemplate.id == template_id,
        WorkScheduleTemplate.gym_id == gym_id
    ).first()

    if not template:
        raise HTTPException(status_code=404, detail="반복 패턴을 찾을 수 없습니다")

    db.delete(template)
    db.commit()

    return None


# ✅ 9. 근무 시간 통계 조회
@router.get("/stats", response_model=List[dict])
def get_work_schedule_stats(
    year_month: str = Query(..., description="YYYY-MM 형식"),
    gym_id: Optional[int] = None,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    코치별 월간 근무 시간 통계
    """
    assert_work_schedule_read_access(current_user)
    gym_id = resolve_accessible_gym_id(db, current_user, gym_id)

    # year_month 파싱
    try:
        year, month = map(int, year_month.split('-'))
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM")

    # 해당 월의 근무 기록 조회
    schedules = db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.date >= first_day,
        WorkSchedule.date <= last_day
    ).all()

    # 코치별로 통계 계산
    stats = {}
    for schedule in schedules:
        if schedule.coach_id not in stats:
            coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
            stats[schedule.coach_id] = {
                'coach_id': schedule.coach_id,
                'coach_name': coach.name if coach else 'Unknown',
                'hourly_wage': coach.hourly_wage if coach else 0,
                'class_wage': coach.class_wage if coach else 0,
                'total_hours': 0.0,
                'total_days': 0,
                'total_classes': 0,
                'expected_salary': 0,
                'shift_details': {'regular': 0, 'overtime': 0, 'holiday': 0}
            }

        # 카테고리별 계산
        if getattr(schedule, 'work_category', 'general') == 'class':
            # 수업인 경우: 횟수 +1, 수업 수당 합산
            stats[schedule.coach_id]['total_classes'] += 1
            class_wage = stats[schedule.coach_id]['class_wage']
            stats[schedule.coach_id]['expected_salary'] += class_wage
        else:
            # 일반 근무인 경우: 근무 시간(hours) 합산, 시급 합산
            start = datetime.strptime(schedule.start_time, "%H:%M").time()
            end = datetime.strptime(schedule.end_time, "%H:%M").time()
            hours = (datetime.combine(date.today(), end) - datetime.combine(date.today(), start)).total_seconds() / 3600

            stats[schedule.coach_id]['total_hours'] += hours
            stats[schedule.coach_id]['total_days'] += 1
            stats[schedule.coach_id]['shift_details'][schedule.shift_type] += 1
            
            hourly_wage = stats[schedule.coach_id]['hourly_wage']
            stats[schedule.coach_id]['expected_salary'] += int(hours * hourly_wage)

    return list(stats.values())


# ✅ 10. 근무표 엑셀 내보내기
@router.get("/export")
def export_work_schedule_excel(
    year_month: str = Query(..., description="YYYY-MM 형식"),
    gym_id: Optional[int] = None,
    current_user: Member = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    근무표를 엑셀 파일로 내보내기
    """
    assert_work_schedule_read_access(current_user)
    gym_id = resolve_accessible_gym_id(db, current_user, gym_id)

    # year_month 파싱
    try:
        year, month = map(int, year_month.split('-'))
        first_day = date(year, month, 1)
        if month == 12:
            last_day = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = date(year, month + 1, 1) - timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM")

    # 지점 정보 조회
    gym = db.query(Gym).filter(Gym.id == gym_id).first()
    gym_name = gym.name if gym else "Unknown"

    # 근무 기록 조회
    schedules = db.query(WorkSchedule).filter(
        WorkSchedule.gym_id == gym_id,
        WorkSchedule.date >= first_day,
        WorkSchedule.date <= last_day
    ).order_by(WorkSchedule.date, WorkSchedule.coach_id).all()

    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "근무표"

    # 스타일 설정
    header_fill = PatternFill(start_color="3182F6", end_color="3182F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 헤더 작성
    headers = ["날짜", "요일", "코치명", "시작시간", "종료시간", "근무유형", "시간", "메모"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # 데이터 작성
    row = 2
    for schedule in schedules:
        coach = db.query(Member).filter(Member.id == schedule.coach_id).first()
        coach_name = coach.name if coach else "Unknown"

        start = datetime.strptime(schedule.start_time, "%H:%M").time()
        end = datetime.strptime(schedule.end_time, "%H:%M").time()
        hours = (datetime.combine(date.today(), end) - datetime.combine(date.today(), start)).total_seconds() / 3600

        shift_label = {
            'regular': '정규',
            'overtime': '연장',
            'holiday': '휴일'
        }.get(schedule.shift_type, schedule.shift_type)

        weekday_name = {
                0: '월', 1: '화', 2: '수', 3: '목',
                4: '금', 5: '토', 6: '일'
            }.get(schedule.date.weekday(), '')

        data = [
            schedule.date.strftime('%Y-%m-%d'),
            weekday_name,
            coach_name,
            schedule.start_time,
            schedule.end_time,
            shift_label,
            f"{hours:.1f}",
            schedule.memo or ""
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col in [1, 7]:  # 날짜, 시간 열 가운데 정렬
                cell.alignment = center_alignment

        row += 1

    # 열 너비 설정
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 20

    # 엑셀 파일을 메모리에 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=근무표_{gym_name}_{year_month}.xlsx"}
    )
